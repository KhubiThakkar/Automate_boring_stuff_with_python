import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

#getting cells from sheets
wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb['Sheet1']
# print(sheet.cell(row = 1, column = 2))
# print(sheet.cell(row=1,column=2).value)
# for i in range(1,8,2):
#     print(i,sheet.cell(row=i,column=2).value)

# #Finding max Row and Column
# print(sheet.max_column,sheet.max_row)

# #Converting between Column numbers and letters
# print(get_column_letter(25))
# print(get_column_letter(600))
# print(column_index_from_string('GH'))
# print(column_index_from_string('RE'))

# #getting rows and columns from sheets

# sheet = wb['Sheet1']
# tuple(sheet['A1':'C3']) #getting all cells from A1 to C3
# for rowOfCellObjects in sheet['A1':'C3']:
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate,cellObj.value)
#     print('---END OF ROW---')

# sheet = wb.active
# list(sheet.columns)[1]
# for cellObj in list(sheet.columns)[1]:
#     print(cellObj.value)
