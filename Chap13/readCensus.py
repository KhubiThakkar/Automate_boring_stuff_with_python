#! python3
# readCensus.py - Tabulates population and number of census tracts for each country

import openpyxl, pprint
print('Opening workbook...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']
countyData={}

#fill in countyData with each county's population and tracts
print('Reading rows...')
for row in range(2,sheet.max_row+1):
    state = sheet['B'+str(row)].value
    county = sheet['C'+str(row)].value
    pop = sheet['D'+str(row)].value

    countyData.setdefault(state,{})
    countyData[state].setdefault(county,{'tracts':0,'pop':0})
    countyData[state][county]['tracts']+=1
    countyData[state][county]['pop'] += int(pop)

print('writing Results...')
resultFile = open('census2010.py','w')
resultFile.write('allData= '+pprint.pformat(countyData))
resultFile.close()
print('Done.')


#reading data of a state
stateData={}
for row in range(2,sheet.max_row+1):
    state = sheet['B'+str(row)].value
    county = sheet['C'+str(row)].value
    stateData.setdefault(state,0)
    stateData[state] = stateData[state]+countyData[state][county]['pop']

print('writing state results...')
resultFile = open('stateCensus.py','w')
resultFile.write('all data = '+pprint.pformat(stateData))
resultFile.close()
print('Done!!')